import asyncio
import discord
import random
import time
import openpyxl
import os
from discord.ext import commands
a=[1] #치이 결혼 변화 수치
반복=[1]
시작=[0] #이거 켜지면 맨션함
대사변수 = random.randrange(1,2) #대사 변수 갯수 (embed 사용)
랜덤대사변수 =random.randrange(1,4) #랜덤 대사 변수
일반대사 = ["아우우우..쉬고있는거에요","아우우우? 호랑이님과 공부하는거에요!"] #대사임
W일반대사 = ["오빠! 지금 쉬고있는거에요! 빨리 놀아주는거에요!","호랑이님이랑 공부하는중인거에요! 끝나면 같이 놀아주는거에요!",] #결혼한후 대사
타이머 = ["아우우우..이제 일어나시는거에요 오라버니!","잠꾸러기 오라버니! 일어나시는거에요!"]
W타이머 = ["오빠! 이제 일어나는거에요!","아우우우...오빠의 자는모습도귀여.....오빠?..일어나있었던건가요?....꺄우우우우!"]
W놀아줘 = ["오빠! 아직 공부중인거에요!" , "아직은 못노는거에요 오빠!"]
놀아줘 = ["아우우우...아직 공부중인거에요..", "아우우우 오라버니 지금은 못노는거에요..아우우우"]
# embed.set_thumbnail(url="https://i.imgur.com/ikAsnSP.jpg")

# a=1은 결혼값 a=0은 결혼값

app = commands.Bot(command_prefix=".")

token = "NTUyODI2MjY4MTkwNjM4MDkw.D2FLjA.JB2ch8EeBrH4ue2PH49EsLGlA1E"
@app.event #치이작동!
async def on_ready():
    print("다음으로 로그인합니다 : ")
    print(app.user.name)
    print(app.user.id)
    print("===============")
    await app.change_presence(game=discord.Game(name="아우우우 치이인거에요!" , type=1))

@app.event #대사 
async def on_message(message):
    
        
    if message.author.bot:
        return None
    
    elif message.content == "치이":
        await message.channel.send("아우우우 부르신건가요?")
        msg = await app.wait_for_message(timeout=15.0)

        if msg is None:
            await message.channel.send("아우우우? 아무것도 아닌가요?")
        elif msg == "놀아줘":
            if a[0]==1:
                embed=discord.Embed(title=random.choice(W놀아줘), color=0x5998dd)
                await message.channel.send(embed=embed)
            elif a[0]==0:
                embed=discord.Embed(title=random.choice(놀아줘), color=0x5998dd)
                await message.channel.send(embed=embed)

    elif message.content.startswith("치이야 뭐해"):
        if a[0]==1:
            embed=discord.Embed(title=random.choice(일반대사), color=0x5998dd)
            await message.channel.send(embed=embed)
        elif a[0]==0:
            embed=discord.Embed(title=random.choice(W일반대사), color=0x5998dd)
            await message.channel.send(embed=embed)

    if message.content.startswith('!타이머'):

        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]

        sec = int(Text)

        for i in range(sec, 0, -1):
            print(i)
            time.sleep(1)
        else:
            if a[0]==0:
                embed=discord.Embed(title=random.choice(W타이머) ,color=0x5998dd)
                await message.channel.send(embed=embed)
    
    if message.content == "!결혼":
        a[0]=0
        embed=discord.Embed(title="아..아우우우?오..오빠?",description="치이와 결혼했습니다 부러운녀석",color=0x5998dd)
        embed.set_thumbnail(url="https://i.imgur.com/ikAsnSP.jpg")
        await message.channel.send(embed=embed)
        
    if message.content == "혼인":
        file = openpyxl.load_workbook("치이결혼.xlsx")
        sheet = file.active
        for i in range(1, 120):
            if int(sheet["B" +str(i)].value) == 1:
                embed=discord.Embed(title="오빠! 이미 혼인한거에요!",color=0x5998dd)
                await message.channel.send(embed=embed)
                break
            elif int(sheet["B" +str(i)].value) != 1:
                if 반복[0] == 1:
                    sheet["B" +str(i)].value = int(sheet["B" +str(i)].value) + 1
                    embed=discord.Embed(title="아..아우우우?오..오빠?",description="치이와 결혼했습니다 ~~부러운녀석~~",color=0x5998dd)
                    embed.set_thumbnail(url="https://i.imgur.com/ikAsnSP.jpg")
                    await message.channel.send(embed=embed)
                    break
            if (sheet["A" + str(i)].value) == "-":
                sheet["A" +str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("치이결혼.xlsx")

    if message.content == '이혼':
        file = openpyxl.load_workbook("치이결혼.xlsx")
        sheet = file.active
        for i in range(1, 120):
            if int(sheet["B" +str(i)].value) == 0:
                    embed=discord.Embed(title="아우우? 혼...인을했던가요??",color=0x5998dd)
                    await message.channel.send(embed=embed)
                    break
            elif int(sheet["B" +str(i)].value) != 0:
                    sheet["B" +str(i)].value = int(sheet["B" +str(i)].value) - 1
                    embed=discord.Embed(title="아..아우우우..오...오라버니..",description="치이와 이혼했습니다 ~~죽고싶냐~~",color=0x5998dd)
                    embed.set_thumbnail(url="https://i.imgur.com/ikAsnSP.jpg")
                    await message.channel.send(embed=embed)
                    break
            if (sheet["A" + str(i)].value) == "-":
                    sheet["A" +str(i)].value = str(message.author.id)
                    sheet["B" + str(i)].value = 0
                    break
        file.save("치이결혼.xlsx")

    if message.content == '!이혼':
        a[0]=1
        embed=discord.Embed(title="아..아우우우..오...오라버니..",description="치이와 이혼했습니다 ~~죽고싶냐~~",color=0x5998dd)
        embed.set_thumbnail(url="https://i.imgur.com/ikAsnSP.jpg")
        await message.channel.send(embed=embed)

    if "궯" in message.content and "밥보" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 51):
            if str(sheet["A" +str(i)].value) == str(message.author.id):
                sheet["B" +str(i)].value = int(sheet["B" +str(i)].value) + 1
                if int(sheet["B" +str(i)].value) == 999:
                    await app.ban(message.author, 1) 
                break
            if (sheet["A" + str(i)].value) == "-":
                sheet["A" +str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")
        await message.channel.send("경고접수!")
access_token = os.environ["BOT_TOKEN"]
app.run(access_token)
