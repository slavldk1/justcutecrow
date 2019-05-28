import asyncio
import discord
import random
import time
import openpyxl
from discord.ext import commands
import os
app = discord.Client()
bot = commands.Bot(command_prefix=".")


@app.event
async def on_ready():
    print("다음으로 로그인합니다")
    print("[폐이임]")
    print("===============")

@app.event
async def on_message(message):
    if message.author.bot:
        return None

    if message.content.startswith("폐이"):
        file = openpyxl.load_workbook("폐이.xlsx")
        sheet = file.active
        for i in range(1,41):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                if sheet["B" + str(i)].value == 1:
                    await message.channel.send("[ㅇ? 나필요해? 왜?]")
                if sheet["B" + str(i)].value == 2:
                    await message.channel.send("[왜? 오빠? 나 필요해?]")
                if sheet["B" + str(i)].value == 0:
                    await message.channel.send("[ㅇ? 나필요해? 왜?]")
                
    if message.content == "체크":
        await message.channel.send("체크됨!")

    if message.content == "결혼":
        file = openpyxl.load_workbook("폐이.xlsx")
        sheet = file.active
        for i in range(1,41):
            if sheet["B" + str(i)].value == 2:
                await message.channel.send("[오빠? 결혼2번할꺼야?]")
                break
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = 2
                await message.channel.send("[???????? 뭐임 진짜임?]")
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 2
                await message.channel.send("[???????? 뭐임 진짜임?]")
                break
        file.save("폐이.xlsx")
        
        

    if message.content == "이혼":
        file = openpyxl.load_workbook("폐이.xlsx")
        sheet = file.active
        for i in range(1,41):
            if sheet["B" + str(i)].value == 1:
                await message.channel.send("[? 우리결혼한적있음?]")
                break
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = 1
                await message.channel.send("....나..울꺼야")
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                await message.channel.send("[? 우리결혼한적있음?]")
                break
        file.save("폐이.xlsx")

        
    
    if message.content == "!폐이야":
        randomNum = random.randrange(1,3)
        if randomNum==1:
            await message.channel.send("[? 나부름? 무슨일?]")
        if randomNum==2:
            await message.channel.send("[내가 그렇게 보고싶었음?]")

app.run("NTUzNTU3NDExNDUzOTkyOTYx.D2P0QA.4Vvd7X9sCVptThCX4Jn29OAQHpQ")
